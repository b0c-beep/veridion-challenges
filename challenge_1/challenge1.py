import pandas as pd
import requests
from bs4 import BeautifulSoup
from termcolor import colored
from colorama import init
import usaddress
import pyap
from address_parser import Parser
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment
import os
from cleantext import clean
from datetime import datetime
import usaddress
from urllib.parse import urljoin, urlparse
from requests.exceptions import RequestException, SSLError
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

# Initialize colorama
init()  

websites_path = './challenge_1/list_of_company_websites.snappy.parquet' # Path to the Parquet file containing the list of websites

# Set display options to show all rows and columns
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)

# Read the Parquet file
df = pd.read_parquet(websites_path, engine='pyarrow') 

# Initialize the geolocator
geolocator = Nominatim(user_agent="address_validator")

def scrape_page_content(url):
    """Fetches and cleans the page content."""
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36'}
        response = requests.get(url, verify=True, headers=headers)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        return soup
    except SSLError:
        # Skip if SSL certificate is invalid
        print(f"Skipping insecure site (SSL verification failed): {url}")
    except RequestException as e:
        # Catch any other request-related errors
        print(f"Error fetching {url}: {e}")
    return None

def scrape_links_and_content(start_url):
    """Scrapes all reachable pages starting from the given URL and visits unique links only."""
    
    try:
        visited_urls = set()  # Set to keep track of visited URLs
        visited_hrefs = set()  # Set to keep track of visited hrefs
        all_scraped_content = ""  # String to hold all scraped content

        page_soup = scrape_page_content(start_url) # Scrape the content of the start URL
        if page_soup is None:
            print(f"Failed to scrape the content of {start_url}")
            return None
        all_scraped_content += page_soup.get_text() + '\n'  # Append the content to the main string

        # Find all <a> tags on the page
        links = page_soup.find_all('a', href=True)  # Only get links that have an href attribute

        # Loop through each link and scrape its content
        for link in links:
            href = link['href']
            full_url = urljoin(start_url, href)  # Construct absolute URL

            # Check if the href is internal and has not been visited
            if urlparse(full_url).netloc == urlparse(start_url).netloc and href not in visited_hrefs:
                visited_urls.add(full_url)  # Mark the URL as visited
                visited_hrefs.add(href)  # Mark the href as visited
                link_soup = scrape_page_content(full_url)  # Scrape the content of the link
                all_scraped_content += link_soup.get_text() + '\n'  # Append the content to the main string

        return all_scraped_content
    except Exception as e:
        print(f"Error scraping links and content: {e}")
        return None

def clean_website_content(text):
    """Clean website content by removing extra spaces and normalizing text."""
    # Remove extra whitespace
    if text is None:
        return text
    cleaned_text = ' '.join(text.split())
    
    # Optionally remove non-alphanumeric characters, keeping basic punctuation
    cleaned_text = ''.join(c for c in cleaned_text if c.isalnum() or c.isspace() or c in ",.;:!?")
    
    return cleaned_text


def extract_pyap(text):
    """Extract both US and UK addresses using pyap."""
    # Regular expressions for detecting both US and UK addresses
    if text is None: 
        return None
    usa_addresses = pyap.parse(text, country='US')
    uk_addresses = pyap.parse(text, country='GB')

    # Convert extracted addresses into strings for easy handling
    extracted_addresses = []

    for address in usa_addresses:
        extracted_addresses.append(str(address))

    for address in uk_addresses:
        extracted_addresses.append(str(address))

    return extracted_addresses if extracted_addresses else None


def count_domains_in_snappy(file_path):
    """Return the number of domains in a Snappy Parquet file."""
    # Read the Parquet file
    df = pd.read_parquet(file_path, engine='pyarrow') 
    
    # Return the number of rows in the 'domain' column
    return df['domain'].nunique()  # Use nunique() if domains might repeat, otherwise use len(df)

def apply_stats_colors(worksheet, stats_df, colors):
    """Apply background colors to stats rows based on the metric."""
    # Apply fill colors to rows based on the metric
    for row_num in range(len(stats_df)):

        if row_num == 0:  
            continue  # Skip the first row (header)
        elif row_num == 1:
            fill_color = colors['Reachable']
        elif row_num == 2:
            fill_color = colors['Unreachable']
        elif row_num == 3:
            fill_color = colors['No Address']
        else:
            fill_color = None

        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            # Apply the fill to the correct row in the worksheet
            for col_num in range(1, len(stats_df.columns) + 1):
                # Adjust for where stats_df starts in the worksheet
                worksheet[f'{get_column_letter(col_num)}{len(results) + row_num + 3}'].fill = fill  # +3 for offset

    
# Save the results to an Excel file
def save_results_to_excel(results):
    """Save the results to an Excel file, adjusting column widths and applying conditional row coloring."""
    df = pd.DataFrame(results)

    output_path = './challenge_1/results.xlsx'

    total_sites = len(results)
    
    # Count reachable, unreachable, reachable-but-no-address and validated sites
    reachable_count = sum(1 for result in results if result['Status'] == 'Reachable')
    unreachable_count = sum(1 for result in results if result['Status'] == 'Unreachable')
    no_address_count = sum(1 for result in results if result['Status'] == 'Reachable - No Addresses')
    validated_count = sum(1 for result in results if result['Status'] == 'Reachable' and result['Validated with GeoPy'] != 'Not validated')

    # Calculate percentages
    reachable_percentage = (reachable_count / total_sites) * 100
    unreachable_percentage = (unreachable_count / total_sites) * 100
    no_address_percentage = (no_address_count / total_sites) * 100
    validated_percentage = (validated_count / total_sites) * 100

    # Create a DataFrame for the stats
    stats_data = {
        'Metric': ['Reachable sites', 'Unreachable sites', 'Reachable but no addresses', 'Reachable with validated address', 'Total sites checked'],
        'Count': [reachable_count, unreachable_count, no_address_count, validated_count, total_sites],
        'Percentage': [f'{reachable_percentage:.2f}%', f'{unreachable_percentage:.2f}%', f'{no_address_percentage:.2f}%', f'{validated_percentage:.2f}%', None]  # Fixed total sites percentage
    }
    stats_df = pd.DataFrame(stats_data)

    # Color for different metrics
    colors = {
        'Reachable': 'C6EFCE',  # Light Green
        'Unreachable': 'FFC7CE',  # Light Red
        'No Address': 'D9D9D9',  # Light Gray
    }

    # Save the results to an Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')

        workbook = writer.book
        worksheet = writer.sheets['Results']

        max_width = 50  # Set max width for columns

        # Adjust column widths
        for col_num, col_name in enumerate(df.columns, start=1):
            max_length = max(df[col_name].astype(str).map(len).max(), len(col_name))
            adjusted_width = min(max_length + 2, max_width)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

            # Ensure text is not wrapped in each cell
            for row_num in range(1, len(df) + 2):  # +2 to account for header
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(wrap_text=False)  # Disable text wrapping

        # Apply conditional coloring to rows based on 'Status'
        for row_num in range(2, len(df) + 2):  # Data starts from row 2 (index 0 in DataFrame)
            status_cell = worksheet[f'C{row_num}']  # Column C holds the 'Status'
            status_value = status_cell.value  # Get the status value from the cell
            
            # Determine the fill color based on the 'Status'
            if status_value == 'Reachable':
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            elif status_value == 'Reachable - No Addresses':
                fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gray (Reachable but no address)
            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red (Unreachable)

            # Apply the fill to the entire row
            for col_num in range(1, len(df.columns) + 1):
                worksheet[f'{get_column_letter(col_num)}{row_num}'].fill = fill
        
        # Write stats DataFrame starting just below the main data
        stats_df.to_excel(writer, index=False, sheet_name='Results', startrow=len(df) + 2)  # Start from the row after the results

        # Call your coloring function here
        apply_stats_colors(worksheet, stats_df, colors)
        

def display_stats(results):
    """Calculate and display statistics of reachable, unreachable, reachable-but-no-address and validated sites."""
    total_sites = len(results)
    
    # Count reachable, unreachable, and reachable-but-no-address
    reachable_count = sum(1 for result in results if result['Status'] == 'Reachable')
    unreachable_count = sum(1 for result in results if result['Status'] == 'Unreachable')
    no_address_count = sum(1 for result in results if result['Status'] == 'Reachable - No Addresses')
    validated_count = sum(1 for result in results if result['Status'] == 'Reachable' and result['Validated with GeoPy'] != 'Not validated')

    # Calculate percentages
    reachable_percentage = (reachable_count / total_sites) * 100
    unreachable_percentage = (unreachable_count / total_sites) * 100
    no_address_percentage = (no_address_count / total_sites) * 100
    validated_percentage = (validated_count / total_sites) * 100

    # Display stats with colors
    print("\n===== " + colored("Summary of Results", 'cyan') + " =====")
    print(f"Total sites checked: {colored(total_sites, 'white')}")
    print(f"Reachable sites: {colored(reachable_count, 'green')} ({colored(f'{reachable_percentage:.2f}%', 'green')})")
    print(f"Validated addresses: {colored(validated_count, 'green')} ({colored(f'{validated_percentage:.2f}%', 'green')})")
    print(f"Unreachable sites: {colored(unreachable_count, 'red')} ({colored(f'{unreachable_percentage:.2f}%', 'red')})")
    print(f"Reachable but no addresses: {colored(no_address_count, 'yellow')} ({colored(f'{no_address_percentage:.2f}%', 'yellow')})")
    print("==============================\n")

def extract_usaddress(text):
    """Extract addresses using the usaddress library."""
    try:
        if text is None:
            return None

        # Parse the text using usaddress
        parsed_addresses = usaddress.tag(text)
        components = parsed_addresses[0]  
        usaddress_result = (
                    'USA',  # Assuming it's a USA address
                    components.get('StateName', None),  # Region (State)
                    components.get('PlaceName', None),  # City
                    components.get('ZipCode', None),  # Postcode
                    components.get('StreetName', None),  # Road
                    components.get('AddressNumber', None)  # Road number
        )
        # Check if all components are not None
        if all(component is not None for component in usaddress_result):
            return usaddress_result
        else: return None

    except Exception as e:
        print(f"Error extracting address using usaddress: {e}")
        return None
    

def parse_address_for_geopy(address_text):
    """Parse and format address using usaddress."""
    try:
        # Prepare the address for geopy
        parsed_address = usaddress.tag(address_text)[0]
        formatted_address = "{}, {}, {}, {}".format(
            parsed_address.get('AddressNumber', ''),
            parsed_address.get('StreetName', ''),
            parsed_address.get('PlaceName', ''),
            parsed_address.get('StateName', ''),
        )
        return formatted_address
    except usaddress.RepeatedLabelError:
        print("Error parsing address:", address_text)
        return None

def validate_address_with_geopy(formatted_address):
    """Validate address with geopy."""
    try:
        location = geolocator.geocode(formatted_address)
        if location:
            return {"latitude": location.latitude, "longitude": location.longitude, "address": location.address}
        else:
            return None
    except (GeocoderTimedOut, GeocoderServiceError) as e:
        print(f"Geocoding error for {formatted_address}: {e}")
        return None


# Initialize results list
results = []

# Get the total number of domains in the Snappy Parquet file
domains_count = count_domains_in_snappy(websites_path)

# Loop through the websites and check their status
for idx, website in enumerate(df['domain']):
    try:
        print(f"Checking website {idx + 1}/{domains_count}: {website}")
        url = 'http://' + website
        response = requests.get(url, timeout=5)  # timeout after 5 seconds
        
        # Check if the status code is not 200
        if response.status_code != 200:
            url = 'https://' + website
            response = requests.get(url, timeout=5)  # timeout after 5 seconds

        # Check if the website is reachable
        if response.status_code == 200:
            print(colored(f"Website {website} is reachable.", 'green'))
            website_content = scrape_links_and_content(url)

            # Clean the website content
            cleaned_content = clean_website_content(website_content)
            
            # Extract addresses using various methods
            usaddress_results = []
            pyap_results = extract_pyap(cleaned_content)
            if pyap_results:
                for address in pyap_results:
                    usaddress_results.append(extract_usaddress(address))

            # Remove duplicates from usaddress_results
            usaddress_results = list(set(usaddress_results))

            # Check if usaddress_results contains only None objects
            if all(result is None for result in usaddress_results):
                usaddress_results = []

            #Validate addresses using geopy
            geopy_validated_addresses = []
            for address in usaddress_results:
                if address is not None:
                    formatted_address = ', '.join(filter(None, address))
                    parsed_geopy = parse_address_for_geopy(formatted_address)
                    geopy_results = validate_address_with_geopy(parsed_geopy)
                    if geopy_results:
                        geopy_validated_addresses.append(address)

            # Store the results
            results.append({
                'Domain': website,
                'URL': f'=HYPERLINK("{url}", "{url}")',
                'Status': 'Reachable' if pyap_results else 'Reachable - No Addresses',
                'Validated with GeoPy': geopy_validated_addresses if geopy_validated_addresses else 'Not validated'
            })

            # Save the results to Excel
            save_results_to_excel(results)

            # Display the statistics
            display_stats(results)
        else:
            print(colored(f"Website {website} is not reachable. Status code: {response.status_code}", 'red'))
            results.append({
                'Domain': website,
                'URL': f'=HYPERLINK("{url}", "{url}")',
                'Status': 'Unreachable'
            })

            save_results_to_excel(results)
            display_stats(results)
    except requests.exceptions.RequestException as e:
        print(colored(f"Error reaching {website}: {e}", 'red'))
        results.append({
            'Domain': website,
            'URL': f'=HYPERLINK("{url}", "{url}")',
            'Status': 'Unreachable',
        })

        save_results_to_excel(results)
        display_stats(results)

